from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from tkinter import *
from tkinter import ttk, messagebox

root = Tk()
root.geometry("1050x400")
root.configure(bg='lightblue')
root.title('CRUD Gem Rey')


# Basic setup with openpyxl
excel_connection = Workbook()
excel_connection = load_workbook('register.xlsx')
excel_opr = excel_connection.active


# Name Option
name_label = Label(root, text="Name:", font="Arial 15", bg="lightblue")
name_label.grid(row=0, column=0)

name_entry = Entry(root, width=20, font="Arial 15")
name_entry.grid(row=0, column=1, pady=5)


# Program Option
program_label = Label(root, text="Program Enrolled:", font="Arial 15", bg="lightblue")
program_label.grid(row=1, column=0, rowspan=5, pady=5)

program_selected = StringVar()
BSIT = Radiobutton(root, text="BS Information Technology", variable=program_selected, value="BSIT", bg="lightblue")
BSCS = Radiobutton(root, text="BS Computer Science", variable=program_selected, value="BSCS", bg="lightblue")
BSCE = Radiobutton(root, text="BS Computer Engineering", variable=program_selected, value="BSCE", bg="lightblue")
BSIS = Radiobutton(root, text="BS Information System", variable=program_selected, value="BSIS", bg="lightblue")

BSIT.grid(row=2, column=1, pady=2, sticky='w')
BSCS.grid(row=3, column=1, pady=2, sticky='w')
BSCE.grid(row=4, column=1, pady=2, sticky='w')
BSIS.grid(row=5, column=1, pady=2, sticky='w')


# Address Option
address_label = Label(root, text="Residence:", font="Arial 15", bg="lightblue")
address_label.grid(row=6, column=0, pady=4)

address = Text(root, width=28, height=3)
address.grid(row=6, column=1, pady=4, sticky='w')


# Submit
submit = Button(root, text="Submit Record", width=20, height=2, font="Arial 10", command=lambda:submit_data())
submit.grid(row=7, column=1, sticky='w')

search = Button(root, text="Search Record", width=20, height=2, font="Arial 10", command=lambda:search_data())
search.grid(row=8, column=1, sticky='w')

delete = Button(root, text="Delete Record", width=20, height=2, font="Arial 10", command=lambda:delete_data())
delete.grid(row=9, column=1, sticky='w')

update = Button(root, text="Update Record", width=20, height=2, font="Arial 10", command=lambda:refresh_tree())
update.grid(row=10, column=1, sticky='w')


tree = ttk.Treeview(root, columns=("NAME", "PROGRAM", "ADDRESS"), show="headings")
tree.heading("NAME", text="NAME")
tree.heading("PROGRAM", text='PROGRAM')
tree.heading("ADDRESS", text="ADDRESS")
tree.grid(row=0, column=3, rowspan=10, padx=10)


def submit_data():
    global name_text
    global program_text
    global address_text
    
    name_text = name_entry.get()
    program_text = program_selected.get()
    address_text = address.get("1.0", "end-1c")

    last_row = str(excel_opr.max_row + 1)
    excel_opr[f"A{last_row}"] = name_text
    excel_opr[f"B{last_row}"] = program_text
    excel_opr[f"C{last_row}"] = address_text

    excel_connection.save('register.xlsx')
    msg = f"DATA ADDED SUCCESSFULLY!"
    messagebox.showinfo("Record", msg)
    tree.insert("", "end", values=(name_text, program_text, address_text))


def view_data():
    for i in range(1, (excel_opr.max_row) + 1):
        tree.insert("", "end", values=(excel_opr["A"+str(i)].value, excel_opr["B"+str(i)].value, excel_opr["C"+str(i)].value))


def search_data():
    global cell_address
    name_text = name_entry.get()
    found = False
    for i in range(1, excel_opr.max_row + 1):
        if name_text in excel_opr["A" + str(i)].value:
            found = True
            cell_address = str(i)
            break
    if found == True:
        name_entry.insert(0, excel_opr["A" + cell_address].value)
        program_value = excel_opr["B" + cell_address].value
        if program_value == "BSIT":
            BSIT.select()
        elif program_value == "BSCS":
            BSCS.select()
        elif program_value == "BSCE":
            BSCE.select()
        elif program_value == "BSIS":
            BSIS.select()
        address.insert("1.0", excel_opr["C" + cell_address].value)


def delete_data():
    excel_opr.delete_rows(int(cell_address))
    messagebox.showerror("DATA DELETED", "DATA REMOVED FROM DB")
    excel_connection.save("register.xlsx")
    tree.insert("", "end", values=(name_text, program_text, address_text))


def get_updated_data():
    update_values = list()
    for i in range(1, excel_opr.max_row + 1):
        update_values.append([excel_opr["A"+str(i)].value, excel_opr["B"+str(i)].value, excel_opr["C"+str(i)].value])
    return update_values

def refresh_tree():
    tree.delete(*tree.get_children())
    data = get_updated_data()
    for item in data:
        tree.insert("", "end", values=item)


view_data()

root.mainloop()